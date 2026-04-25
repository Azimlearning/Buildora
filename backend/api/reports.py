from fastapi import APIRouter, HTTPException
from fastapi.responses import Response
from pydantic import BaseModel
from backend.core.firebase_client import get_firestore_client
import json
import os
import io
import datetime
from fpdf import FPDF
from openpyxl import Workbook

router = APIRouter()

from typing import Optional, Dict, Any

class CibdReportRequest(BaseModel):
    doc_type: str
    project: Optional[Dict[str, Any]] = None

class ReportGenerateRequest(BaseModel):
    project: Optional[Dict[str, Any]] = None

# In-memory store for generated reports
# In a real app, these would be saved to cloud storage (S3/GCS/Minio)
_GENERATED_REPORTS = {}

@router.get("/reports/{project_id}")
async def get_reports_status(project_id: str):
    """Get the status of main reports (PDF, XLSX)"""
    # Look up from memory
    pdf_status = "queued"
    xlsx_status = "queued"
    
    if project_id in _GENERATED_REPORTS:
        if "pdf" in _GENERATED_REPORTS[project_id]:
            pdf_status = "ready"
        if "xlsx" in _GENERATED_REPORTS[project_id]:
            xlsx_status = "ready"

    return {
        "pdf": {"status": pdf_status, "file_size": "2.1 MB", "page_count": 4, "generated_at": datetime.datetime.now().isoformat()} if pdf_status == "ready" else {"status": pdf_status},
        "xlsx": {"status": xlsx_status, "file_size": "150 KB", "row_count": 45, "generated_at": datetime.datetime.now().isoformat()} if xlsx_status == "ready" else {"status": xlsx_status}
    }

@router.post("/reports/{project_id}/generate")
async def generate_report(project_id: str, req: ReportGenerateRequest = None):
    """Generate main PDF and XLSX reports"""
    db = get_firestore_client()
    
    project = req.project if req and req.project else None
    if not project:
        project = await db.get_project(project_id)
    
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    if project_id not in _GENERATED_REPORTS:
        _GENERATED_REPORTS[project_id] = {}

    def clean_text(text):
        if not text: return "N/A"
        return str(text).encode('latin-1', 'replace').decode('latin-1')

    # Generate PDF
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=16)
    pdf.cell(200, 10, txt=f"Project Report: {clean_text(project.get('name', 'Unknown'))}", ln=1, align='C')
    pdf.set_font("Arial", size=12)
    pdf.ln(10)
    pdf.cell(200, 10, txt=f"Contractor: {clean_text(project.get('contractor', 'Unknown'))}", ln=1)
    pdf.cell(200, 10, txt=f"CIDB Grade: {clean_text(project.get('cidb_grade', 'N/A'))}", ln=1)
    pdf.cell(200, 10, txt=f"Health Score: {clean_text(project.get('health_score', 'N/A'))}%", ln=1)
    
    pdf.ln(10)
    pdf.cell(200, 10, txt="Summary:", ln=1)
    pdf.set_font("Arial", size=10)
    pdf.multi_cell(0, 10, txt=clean_text(project.get('description', 'No description available.')))
    
    out = pdf.output(dest='S')
    _GENERATED_REPORTS[project_id]["pdf"] = bytes(out) if isinstance(out, (bytearray, bytes)) else out.encode('latin1')

    # Generate XLSX
    wb = Workbook()
    ws = wb.active
    ws.title = "Project Tracker"
    ws.append(["Project Name", "Contractor", "CIDB Grade", "Health Score"])
    ws.append([project.get('name', 'Unknown'), project.get('contractor', 'Unknown'), project.get('cidb_grade', 'N/A'), project.get('health_score', 'N/A')])
    
    ws.append([])
    ws.append(["Milestone", "Status", "Variance"])
    ws.append(["Foundation", "Complete", "0%"])
    ws.append(["Structure", "In Progress", "+5%"])
    ws.append(["Finishes", "Pending", "0%"])
    
    excel_io = io.BytesIO()
    wb.save(excel_io)
    excel_io.seek(0)
    _GENERATED_REPORTS[project_id]["xlsx"] = excel_io.read()

    return {"status": "success", "message": "Report generation triggered"}

@router.get("/reports/{project_id}/download/{format}")
async def download_report(project_id: str, format: str):
    """Download generated report"""
    if project_id not in _GENERATED_REPORTS or format not in _GENERATED_REPORTS[project_id]:
        raise HTTPException(status_code=404, detail="Report not generated yet")
    
    data = _GENERATED_REPORTS[project_id][format]
    
    if format == "pdf":
        return Response(content=data, media_type="application/pdf")
    elif format == "xlsx":
        return Response(content=data, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        raise HTTPException(status_code=400, detail="Invalid format")

@router.post("/reports/{project_id}/generate/cibd")
async def generate_cibd_report(project_id: str, req: CibdReportRequest):
    """Generate CIBD specific reports from knowledge base"""
    db = get_firestore_client()
    
    project = req.project if req and req.project else None
    if not project:
        project = await db.get_project(project_id)
    
    # Load knowledge base
    try:
        with open("knowledge_base.json", "r") as f:
            kb = json.load(f)
    except:
        kb = {}
        
    if project_id not in _GENERATED_REPORTS:
        _GENERATED_REPORTS[project_id] = {}
        
    doc_type = req.doc_type
    
    def clean_text(text):
        if not text: return "N/A"
        return str(text).encode('latin-1', 'replace').decode('latin-1')

    # Generate a simple PDF for the CIBD document
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", 'B', size=16)
    title = doc_type.replace('_', ' ').title()
    pdf.cell(200, 10, txt=f"CIDB Form: {clean_text(title)}", ln=1, align='C')
    
    pdf.set_font("Arial", size=12)
    pdf.ln(10)
    pdf.cell(200, 10, txt=f"Project: {clean_text(project.get('name', 'Unknown') if project else 'Unknown')}", ln=1)
    pdf.cell(200, 10, txt=f"Date Generated: {datetime.datetime.now().strftime('%Y-%m-%d')}", ln=1)
    
    pdf.ln(10)
    pdf.set_font("Arial", 'B', size=12)
    pdf.cell(200, 10, txt="Requirements from Knowledge Base:", ln=1)
    pdf.set_font("Arial", size=10)
    
    # Try to find requirements in KB
    doc_reqs = kb.get("requirements", {}).get("documents", {}).get(doc_type, {})
    if not doc_reqs:
        # Fallback to general documents list
        doc_reqs = kb.get("documents", {})
        
    for k, v in doc_reqs.items():
        if isinstance(v, dict):
            pdf.multi_cell(0, 10, txt=clean_text(f"- {k.replace('_', ' ').title()}: {v.get('description', 'Required')}"))
        else:
            pdf.cell(200, 10, txt=clean_text(f"- {k.replace('_', ' ').title()}"), ln=1)
            
    # Save to memory
    out = pdf.output(dest='S')
    _GENERATED_REPORTS[project_id][f"cibd_{doc_type}"] = bytes(out) if isinstance(out, (bytearray, bytes)) else out.encode('latin1')
    
    return {"status": "success", "message": f"{doc_type} generated"}

@router.get("/reports/{project_id}/download/cibd/{doc_type}")
async def download_cibd_report(project_id: str, doc_type: str):
    """Download CIBD report"""
    key = f"cibd_{doc_type}"
    if project_id not in _GENERATED_REPORTS or key not in _GENERATED_REPORTS[project_id]:
        raise HTTPException(status_code=404, detail="Report not generated yet")
        
    data = _GENERATED_REPORTS[project_id][key]
    return Response(content=data, media_type="application/pdf")
