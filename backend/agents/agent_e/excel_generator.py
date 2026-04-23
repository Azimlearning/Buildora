"""
Excel Report Generator
======================
Generates XLSX cost tracker and project reports using openpyxl.

Reports include:
  - Cost breakdown and variance analysis
  - Timeline tracking
  - Budget vs actual comparison
  - Milestone tracking
"""

from typing import Dict, Any, List
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelGenerator:
    """XLSX report generator using openpyxl."""

    def __init__(self):
        self.header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    async def generate_cost_tracker(
        self,
        project_id: str,
        project_data: Dict[str, Any]
    ) -> str:
        """
        Generate XLSX cost tracker with budget breakdown.

        Args:
            project_id:   Project identifier
            project_data: Complete project data from Firestore

        Returns:
            Path to generated XLSX file
        """
        output_path = f"/tmp/cost_tracker_{project_id}.xlsx"
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create sheets
        self._create_summary_sheet(wb, project_data)
        self._create_cost_breakdown_sheet(wb, project_data)
        self._create_timeline_sheet(wb, project_data)

        # Save workbook
        wb.save(output_path)
        return output_path

    def _create_summary_sheet(self, wb: Workbook, project_data: Dict[str, Any]):
        """Create project summary sheet."""
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws['A1'] = "Buildora Cost Tracker"
        ws['A1'].font = Font(bold=True, size=16, color="1a73e8")
        ws.merge_cells('A1:D1')

        # Project info
        project = project_data.get('project', {})
        ws['A3'] = "Project Information"
        ws['A3'].font = self.header_font
        ws['A3'].fill = self.header_fill

        info_data = [
            ['Project ID', project.get('id', 'N/A')],
            ['Project Name', project.get('name', 'N/A')],
            ['Status', project.get('status', 'N/A')],
            ['Generated', datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')],
        ]

        row = 4
        for label, value in info_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

        # Compliance summary
        ws[f'A{row + 1}'] = "Compliance Summary"
        ws[f'A{row + 1}'].font = self.header_font
        ws[f'A{row + 1}'].fill = self.header_fill

        compliance = project_data.get('compliance_score', {})
        compliance_data = [
            ['Compliance Score', f"{compliance.get('score', 0):.1f}%"],
            ['Status', compliance.get('status', 'unknown').upper()],
            ['Missing Documents', len(compliance.get('gaps', []))],
        ]

        row += 2
        for label, value in compliance_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40

    def _create_cost_breakdown_sheet(self, wb: Workbook, project_data: Dict[str, Any]):
        """Create cost breakdown sheet."""
        ws = wb.create_sheet("Cost Breakdown")

        # Headers
        headers = ['Category', 'Budgeted (RM)', 'Actual (RM)', 'Variance (RM)', 'Variance (%)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border

        # Sample cost data (to be replaced with real data from extracted fields)
        cost_categories = [
            ['Materials', 150000, 145000, -5000, -3.33],
            ['Labor', 200000, 210000, 10000, 5.00],
            ['Equipment', 80000, 82000, 2000, 2.50],
            ['Permits & Fees', 30000, 28000, -2000, -6.67],
            ['Contingency', 40000, 35000, -5000, -12.50],
        ]

        row = 2
        for category_data in cost_categories:
            for col, value in enumerate(category_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
                if col > 1:  # Numeric columns
                    cell.alignment = Alignment(horizontal='right')
                    if col in [2, 3, 4]:  # Currency columns
                        cell.number_format = '#,##0.00'
                    elif col == 5:  # Percentage column
                        cell.number_format = '0.00%'
                        # Color code variance
                        if value > 0:
                            cell.font = Font(color="FF0000")  # Red for over budget
                        else:
                            cell.font = Font(color="00FF00")  # Green for under budget
            row += 1

        # Total row
        total_row = row
        ws[f'A{total_row}'] = 'TOTAL'
        ws[f'A{total_row}'].font = Font(bold=True)
        ws[f'B{total_row}'] = f'=SUM(B2:B{row-1})'
        ws[f'C{total_row}'] = f'=SUM(C2:C{row-1})'
        ws[f'D{total_row}'] = f'=SUM(D2:D{row-1})'
        ws[f'E{total_row}'] = f'=D{total_row}/B{total_row}'

        for col in range(1, 6):
            cell = ws.cell(row=total_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            cell.border = self.border

        # Auto-adjust column widths
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _create_timeline_sheet(self, wb: Workbook, project_data: Dict[str, Any]):
        """Create timeline tracking sheet."""
        ws = wb.create_sheet("Timeline")

        # Headers
        headers = ['Milestone', 'Planned Date', 'Actual Date', 'Status', 'Delay (Days)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border

        # Sample timeline data (to be replaced with real data)
        milestones = [
            ['Project Kickoff', '2026-01-01', '2026-01-01', 'Completed', 0],
            ['Design Approval', '2026-02-15', '2026-02-20', 'Completed', 5],
            ['Permit Submission', '2026-03-01', '2026-03-05', 'Completed', 4],
            ['Construction Start', '2026-04-01', 'Pending', 'In Progress', 0],
            ['Phase 1 Complete', '2026-06-30', 'Pending', 'Pending', 0],
        ]

        row = 2
        for milestone_data in milestones:
            for col, value in enumerate(milestone_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
                if col in [2, 3]:  # Date columns
                    cell.alignment = Alignment(horizontal='center')
                elif col == 5:  # Delay column
                    cell.alignment = Alignment(horizontal='center')
                    if isinstance(value, (int, float)) and value > 0:
                        cell.font = Font(color="FF0000")  # Red for delays
            row += 1

        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
